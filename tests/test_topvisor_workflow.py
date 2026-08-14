import io
import logging
import urllib.error
from datetime import date
from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model
from django.urls import reverse
from django.utils import timezone
from docx import Document
from openpyxl import load_workbook

from apps.metrics.models import KeywordPosition, RankingSnapshot
from apps.projects.models import Project
from apps.reports.exporting import generate_artifact
from apps.reports.models import Report
from apps.reports.services import create_report_version
from apps.topvisor.client import (
    TopvisorClient,
    TopvisorCredentials,
    TopvisorError,
    TopvisorTemporaryError,
)
from apps.topvisor.models import TopvisorProjectMapping
from apps.topvisor.services import (
    VISIBILITY_FORMULA_URL,
    VISIBILITY_FORMULA_VERSION,
    calculate_visibility,
    sync_positions,
)

pytestmark = pytest.mark.django_db


def test_visibility_uses_exact_official_topvisor_weights():
    rows = [
        {"position": 1, "frequency": 100},
        {"position": 2, "frequency": 200},
        {"position": 10, "frequency": 300},
        {"position": 11, "frequency": 400},
    ]
    # (100*1 + 200*1 + 300*.2 + 400*.1) / 1000 * 100
    assert calculate_visibility(rows) == Decimal("40.0000")


@pytest.mark.parametrize(
    ("position", "expected"),
    [
        (1, "100.0000"),
        (3, "100.0000"),
        (4, "85.0000"),
        (5, "60.0000"),
        (6, "50.0000"),
        (7, "50.0000"),
        (8, "30.0000"),
        (9, "30.0000"),
        (10, "20.0000"),
        (11, "10.0000"),
        (15, "10.0000"),
        (16, "5.0000"),
        (20, "5.0000"),
        (21, "0.0000"),
    ],
)
def test_visibility_official_weight_range_boundaries(position, expected):
    assert calculate_visibility([{"position": position, "frequency": 100}]) == Decimal(expected)


REAL_SEARCH_CONFIGURATION_RESPONSE = [
    {
        "id": 22653133,
        "searchers": [
            {
                "id": 15263163,
                "project_id": 22653133,
                "searcher": 0,
                "enabled": 1,
                "key": 0,
                "name": "Yandex",
                "regions": [
                    {
                        "id": 17054799,
                        "key": 213,
                        "lang": "ru",
                        "device": 0,
                        "depth": 1,
                        "index": 1,
                        "enabled": 1,
                        "searcher_key": 0,
                        "type": "CITY",
                        "countryCode": "RU",
                        "name": "Москва",
                        "areaName": "Москва и Московская область",
                        "domain": ".ru",
                    }
                ],
            },
            {
                "id": 15263165,
                "project_id": 22653133,
                "searcher": 1,
                "enabled": 1,
                "key": 1,
                "name": "Google",
                "regions": [
                    {
                        "id": 17054801,
                        "key": 213,
                        "lang": "ru",
                        "device": 0,
                        "depth": 2,
                        "index": 2,
                        "enabled": 1,
                        "searcher_key": 1,
                        "type": "CITY",
                        "countryCode": "RU",
                        "name": "Москва",
                        "areaName": "Москва и Московская область",
                        "domain": ".ru",
                    }
                ],
            },
        ],
    }
]


def test_search_configurations_use_projects_endpoint_and_flatten_real_response(monkeypatch):
    api = TopvisorClient(credentials=TopvisorCredentials("user", "secret"))
    calls = []

    def request(method, params):
        calls.append((method, params))
        return REAL_SEARCH_CONFIGURATION_RESPONSE

    monkeypatch.setattr(api, "_request", request)
    configurations = api.get_search_configurations(22653133)

    assert calls == [
        (
            "get/projects_2/projects",
            {
                "show_searchers_and_regions": 2,
                "limit": 1,
                "filters": [{"name": "id", "operator": "EQUALS", "values": ["22653133"]}],
            },
        )
    ]
    assert [
        (item["searcher_name"], item["region_name"], item["normalized_depth"])
        for item in configurations
    ] == [
        ("Yandex", "Москва", 100),
        ("Google", "Москва", 20),
    ]
    assert [item["region_index"] for item in configurations] == [1, 2]
    assert configurations[0] == {
        "searcher_id": 15263163,
        "searcher_key": 0,
        "searcher_name": "Yandex",
        "region_id": 17054799,
        "region_key": 213,
        "region_index": 1,
        "region_name": "Москва",
        "area_name": "Москва и Московская область",
        "language": "ru",
        "raw_depth": 1,
        "normalized_depth": 100,
        "device": 0,
    }


@pytest.mark.parametrize(
    "raw, normalized",
    [(1, 10), (2, 20), (3, 30), (5, 50), (10, 100), (20, 20), (30, 30), (50, 50), (100, 100)],
)
def test_topvisor_depth_normalization(raw, normalized):
    assert TopvisorClient._normalize_depth(raw) == normalized


class FakeClient:
    def __init__(self, missing_frequency=False):
        self.calls = []
        self.missing_frequency = missing_frequency

    def get_positions(self, project_id, **filters):
        self.calls.append((project_id, filters))
        row = {"query": "купить слона", "position": 7, "frequency": 120}
        if self.missing_frequency:
            row.pop("frequency")
        return iter([row])


class FailingLaterClient(FakeClient):
    def __init__(self, *, fail_at, secret=""):
        super().__init__()
        self.fail_at = fail_at
        self.secret = secret

    def get_positions(self, project_id, **filters):
        self.calls.append((project_id, filters))
        if len(self.calls) == self.fail_at:
            raise TopvisorError(f"Ошибка провайдера {self.secret}")
        return iter([{"query": "купить слона", "position": 7, "frequency": 120}])


def mapping(project, depth=20, engine="google", region="Москва"):
    return TopvisorProjectMapping.objects.create(
        project=project,
        topvisor_project_id="42",
        topvisor_project_name="Safe project",
        selected_configurations=[
            {
                "id": "7",
                "search_engine": engine,
                "region_name": region,
                "depth": depth,
                "device": "desktop",
            }
        ],
    )


def test_missing_credentials_has_clear_state(client, settings):
    settings.TOPVISOR_USER_ID = settings.TOPVISOR_API_KEY = ""
    user = get_user_model().objects.create_user("user", password="password")
    project = Project.objects.create(name="Site", domain="site.example")
    client.force_login(user)
    response = client.get(reverse("topvisor:connection", args=[project.id]))
    assert response.status_code == 200
    assert "Общие реквизиты Topvisor не настроены" in response.content.decode()
    assert 'value="password"' not in response.content.decode()


def test_connection_and_project_configuration_selection(client, settings, monkeypatch):
    settings.TOPVISOR_USER_ID, settings.TOPVISOR_API_KEY = "uid", "super-secret"
    user = get_user_model().objects.create_user("user", password="password")
    project = Project.objects.create(name="Site", domain="connect.example")
    client.force_login(user)
    monkeypatch.setattr(
        TopvisorClient, "iter_projects", lambda self: iter([{"id": 42, "name": "SEO"}])
    )
    monkeypatch.setattr(
        TopvisorClient,
        "get_search_configurations",
        lambda self, _id: [
            {"id": 7, "search_engine": "google", "region_name": "Москва", "depth": 50}
        ],
    )
    response = client.post(
        reverse("topvisor:connection", args=[project.id]),
        {"topvisor_project": "42", "configurations": ["7"]},
    )
    assert response.status_code == 302
    saved = project.topvisor_mapping
    assert saved.topvisor_project_id == "42"
    assert "super-secret" not in repr(saved.selected_configurations)
    page = client.get(
        reverse("topvisor:connection", args=[project.id]),
        {"topvisor_project": "42"},
    ).content.decode()
    assert 'class="configuration-form"' in page
    assert 'class="configuration-list"' in page


def test_sync_is_idempotent_requires_frequency_and_keeps_depth_per_segment():
    project = Project.objects.create(name="Site", domain="sync.example")
    selected = mapping(project, depth=20)
    selected.selected_configurations.append(
        {"id": "8", "search_engine": "yandex", "region_name": "Россия", "depth": 100}
    )
    selected.save()
    fake = FakeClient()
    first = sync_positions(mapping=selected, report_month=date(2026, 7, 1), client=fake)
    second = sync_positions(mapping=selected, report_month=date(2026, 7, 1), client=fake)
    assert first.status == first.Status.SUCCESS
    assert len(fake.calls) == 12
    assert RankingSnapshot.objects.count() == 6
    assert KeywordPosition.objects.count() == 6
    assert {(x["search_engine"], x["region"], x["depth"]) for x in second.segments} == {
        ("google", "Москва", 20),
        ("yandex", "Россия", 100),
    }
    assert all(
        item.response_checksum
        and item.retrieved_at
        and item.provenance
        and item.visibility is not None
        and item.visibility_raw["source"] == "calculated_from_positions_and_frequency"
        and item.visibility_raw["formula_version"] == VISIBILITY_FORMULA_VERSION
        and item.visibility_raw["formula_source"] == VISIBILITY_FORMULA_URL
        for item in RankingSnapshot.objects.all()
    )
    failed = sync_positions(
        mapping=selected, report_month=date(2026, 8, 1), client=FakeClient(True)
    )
    assert failed.status == failed.Status.FAILED
    assert "частотности" in failed.error_message


def test_late_api_error_leaves_no_partial_snapshots_and_records_failed_run():
    project = Project.objects.create(name="Atomic", domain="atomic.example")
    selected = mapping(project)
    run = sync_positions(
        mapping=selected,
        report_month=date(2026, 7, 1),
        client=FailingLaterClient(fail_at=3),
    )
    assert run.status == run.Status.FAILED
    assert run.completed_at is not None
    assert RankingSnapshot.objects.filter(project=project).count() == 0


def test_late_api_error_does_not_change_existing_snapshots():
    project = Project.objects.create(name="Existing", domain="existing.example")
    selected = mapping(project)
    successful = sync_positions(
        mapping=selected, report_month=date(2026, 7, 1), client=FakeClient()
    )
    assert successful.status == successful.Status.SUCCESS
    before = list(
        RankingSnapshot.objects.filter(project=project)
        .order_by("snapshot_date")
        .values_list("id", "response_checksum", "retrieved_at")
    )
    failed = sync_positions(
        mapping=selected,
        report_month=date(2026, 7, 1),
        client=FailingLaterClient(fail_at=3),
    )
    after = list(
        RankingSnapshot.objects.filter(project=project)
        .order_by("snapshot_date")
        .values_list("id", "response_checksum", "retrieved_at")
    )
    assert failed.status == failed.Status.FAILED
    assert after == before


def test_sync_error_redacts_client_credentials():
    project = Project.objects.create(name="Secret", domain="secret.example")
    selected = mapping(project)
    fake = FailingLaterClient(fail_at=2, secret="api-secret")
    fake.credentials = TopvisorCredentials("user-secret", "api-secret")
    run = sync_positions(mapping=selected, report_month=date(2026, 7, 1), client=fake)
    assert run.status == run.Status.FAILED
    assert "api-secret" not in run.error_message
    assert "[скрыто]" in run.error_message


def test_duplicate_normalized_engine_region_pair_is_rejected(client, settings, monkeypatch):
    settings.TOPVISOR_USER_ID, settings.TOPVISOR_API_KEY = "uid", "key"
    user = get_user_model().objects.create_user("duplicate", password="password")
    project = Project.objects.create(name="Duplicate", domain="duplicate.example")
    client.force_login(user)
    configurations = [
        {
            "id": 7,
            "search_engine": "Google",
            "region_name": "Москва",
            "depth": 50,
            "device": "desktop",
        },
        {
            "id": 8,
            "search_engine": " google ",
            "region_name": "  МОСКВА ",
            "depth": 50,
            "device": "mobile",
        },
    ]
    monkeypatch.setattr(
        TopvisorClient, "iter_projects", lambda self: iter([{"id": 42, "name": "SEO"}])
    )
    monkeypatch.setattr(
        TopvisorClient, "get_search_configurations", lambda self, _id: configurations
    )
    response = client.post(
        reverse("topvisor:connection", args=[project.id]),
        {"topvisor_project": "42", "configurations": ["7", "8"]},
    )
    assert response.status_code == 200
    assert "устройство не является измерением" in response.content.decode()
    assert not TopvisorProjectMapping.objects.filter(project=project).exists()


def test_sync_version_and_docx_xlsx_export_do_not_need_live_api(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    project = Project.objects.create(name="Site", domain="export-sync.example")
    sync_positions(mapping=mapping(project), report_month=date(2026, 7, 1), client=FakeClient())
    version = create_report_version(
        report=Report.objects.create(project=project, report_month=date(2026, 7, 1))
    )
    checksum = version.snapshot.checksum
    docx = generate_artifact(version=version, artifact_type="docx", is_draft=True)
    xlsx = generate_artifact(version=version, artifact_type="xlsx", is_draft=True)
    with docx.file.open("rb") as stream:
        assert Document(io.BytesIO(stream.read())).paragraphs
    with xlsx.file.open("rb") as stream:
        assert "Позиции" in load_workbook(io.BytesIO(stream.read())).sheetnames
    version.snapshot.refresh_from_db()
    assert version.snapshot.checksum == checksum


def test_client_pagination_and_429_retry(monkeypatch):
    responses = [
        urllib.error.HTTPError("url", 429, "rate", {"Retry-After": "0"}, None),
        {"result": {"rows": [{"id": 1}]}},
    ]

    class Response:
        def __enter__(self):
            return self

        def __exit__(self, *args):
            return None

        def read(self):
            return b'{"result":{"rows":[{"id":1}]}}'

    def urlopen(*args, **kwargs):
        value = responses.pop(0)
        if isinstance(value, Exception):
            raise value
        return Response()

    monkeypatch.setattr("urllib.request.urlopen", urlopen)
    sleeps = []
    client = TopvisorClient(
        credentials=TopvisorCredentials("u", "secret"), max_retries=1, sleep=sleeps.append
    )
    assert list(client.iter_pages("method", page_size=2)) == [{"id": 1}]
    assert sleeps == [0.0]


def test_client_retries_api_level_errors_after_an_earlier_success(monkeypatch):
    payloads = [
        b'{"result":{"rows":[{"id":1}]}}',
        b'{"errors":[{"code":429,"message":"secret response body"}]}',
        b'{"result":{"rows":[{"id":2}]}}',
    ]

    class Response:
        def __enter__(self):
            return self

        def __exit__(self, *args):
            return None

        def read(self):
            return payloads.pop(0)

    monkeypatch.setattr("urllib.request.urlopen", lambda *args, **kwargs: Response())
    sleeps = []
    api = TopvisorClient(
        credentials=TopvisorCredentials("user-secret", "api-secret"),
        max_retries=1,
        sleep=sleeps.append,
    )

    assert list(api.iter_projects()) == [{"id": 1}]
    assert list(api.iter_projects()) == [{"id": 2}]
    assert len(sleeps) == 1


def test_permanent_api_level_error_is_not_retried_or_leaked(monkeypatch, caplog):
    calls = 0

    class Response:
        def __enter__(self):
            return self

        def __exit__(self, *args):
            return None

        def read(self):
            return b'{"errors":[{"code":401,"message":"api-secret Authorization"}]}'

    def urlopen(*args, **kwargs):
        nonlocal calls
        calls += 1
        return Response()

    monkeypatch.setattr("urllib.request.urlopen", urlopen)
    api = TopvisorClient(
        credentials=TopvisorCredentials("user-secret", "api-secret"), max_retries=3
    )
    with caplog.at_level(logging.DEBUG), pytest.raises(TopvisorError) as raised:
        api.check_access()

    exposed = str(raised.value) + caplog.text
    assert calls == 1
    assert "api-secret" not in exposed
    assert "user-secret" not in exposed
    assert "Authorization" not in exposed


def test_invalid_credentials_error_never_contains_secret(monkeypatch):
    error = urllib.error.HTTPError("url", 401, "secret-key", {}, None)
    monkeypatch.setattr("urllib.request.urlopen", lambda *a, **k: (_ for _ in ()).throw(error))
    client = TopvisorClient(credentials=TopvisorCredentials("u", "secret-key"), max_retries=0)
    with pytest.raises(TopvisorError) as raised:
        client.check_access()
    assert "secret-key" not in str(raised.value)


def test_global_credentials_are_encrypted_replaced_retained_and_deleted(
    client, settings, monkeypatch
):
    from apps.topvisor.models import TopvisorCredential

    settings.CREDENTIAL_ENCRYPTION_KEY = "stable-test-encryption-key"
    settings.TOPVISOR_USER_ID = settings.TOPVISOR_API_KEY = ""
    user = get_user_model().objects.create_user(
        "credentials", password="login-secret", is_staff=True
    )
    project = Project.objects.create(name="Credentials", domain="credentials.example")
    client.force_login(user)
    monkeypatch.setattr(TopvisorClient, "check_access", lambda self: ({"id": 42, "name": "SEO"},))
    url = reverse("topvisor:credentials")

    response = client.post(
        url, {"action": "credentials", "user_id": "123", "api_key": "first-api-secret"}
    )
    assert response.status_code == 302
    connection = TopvisorCredential.objects.get(pk=1)
    assert b"first-api-secret" not in bytes(connection.api_key_encrypted)
    assert bytes(connection.api_key_encrypted) != b"pending"
    assert connection.get_api_key() == "first-api-secret"
    assert connection.last_verified_at
    html = client.get(url).content.decode()
    assert "first-api-secret" not in html
    assert "Ключ настроен" in html
    project_page = client.get(reverse("topvisor:connection", args=[project.id])).content.decode()
    assert "Проект и поиск" in project_page
    assert "ID пользователя Topvisor" not in project_page

    client.post(url, {"action": "credentials", "user_id": "456", "api_key": ""})
    connection.refresh_from_db()
    assert connection.user_id == "456"
    assert connection.get_api_key() == "first-api-secret"
    client.post(url, {"action": "credentials", "user_id": "456", "api_key": "second-secret"})
    connection.refresh_from_db()
    assert connection.get_api_key() == "second-secret"

    assert client.post(url, {"action": "delete"}).status_code == 302
    assert not TopvisorCredential.objects.exists()
    assert Project.objects.filter(pk=project.pk).exists()


def test_invalid_global_credentials_are_not_saved_or_echoed(client, settings, monkeypatch):
    from apps.topvisor.models import TopvisorCredential

    settings.CREDENTIAL_ENCRYPTION_KEY = "stable-test-encryption-key"
    user = get_user_model().objects.create_user("invalid", password="login-secret", is_staff=True)
    client.force_login(user)
    monkeypatch.setattr(
        TopvisorClient,
        "check_access",
        lambda self: (_ for _ in ()).throw(TopvisorError("response api-secret Authorization")),
    )
    response = client.post(
        reverse("topvisor:credentials"),
        {"action": "credentials", "user_id": "bad", "api_key": "api-secret"},
    )
    html = response.content.decode()
    assert response.status_code == 200
    assert "Не удалось проверить" in html
    assert "api-secret" not in html
    assert "Authorization" not in html
    assert not TopvisorCredential.objects.exists()


def test_credentials_projects_and_configurations_reuse_checked_project_list(
    client, settings, monkeypatch
):
    settings.CREDENTIAL_ENCRYPTION_KEY = "stable-test-encryption-key"
    project = Project.objects.create(name="One request", domain="one-request.example")
    user = get_user_model().objects.create_user("one-request", password="secret", is_staff=True)
    client.force_login(user)
    calls = {"projects": 0, "configurations": 0}

    def projects(api):
        calls["projects"] += 1
        return iter([{"id": 42, "name": "SEO"}])

    def configurations(api, project_id):
        calls["configurations"] += 1
        assert project_id == "42"
        return [{"id": 7, "search_engine": "google", "region_name": "Москва"}]

    monkeypatch.setattr(TopvisorClient, "iter_projects", projects)
    monkeypatch.setattr(TopvisorClient, "get_search_configurations", configurations)
    settings_url = reverse("topvisor:credentials")
    project_url = reverse("topvisor:connection", args=[project.id])

    assert (
        client.post(
            settings_url,
            {"action": "credentials", "user_id": "uid", "api_key": "api-secret"},
        ).status_code
        == 302
    )
    assert client.get(project_url).status_code == 200
    response = client.get(project_url, {"topvisor_project": "42"})

    assert response.status_code == 200
    assert "Москва" in response.content.decode()
    assert calls == {"projects": 1, "configurations": 1}
    assert "api-secret" not in response.content.decode()


def test_temporary_verification_error_preserves_connection_and_mapping(
    client, settings, monkeypatch
):
    project = Project.objects.create(name="Temporary", domain="temporary.example")
    connection = _connection(project, settings)
    selected = mapping(project)
    user = get_user_model().objects.create_user("temporary", password="secret", is_staff=True)
    client.force_login(user)
    monkeypatch.setattr(
        TopvisorClient,
        "check_access",
        lambda self: (_ for _ in ()).throw(TopvisorTemporaryError("safe")),
    )

    response = client.post(
        reverse("topvisor:credentials"),
        {"action": "credentials", "user_id": "replacement", "api_key": "new-secret"},
    )

    connection.refresh_from_db()
    assert response.status_code == 200
    assert "временно недоступен" in response.content.decode()
    assert "new-secret" not in response.content.decode()
    assert connection.user_id == "uid"
    assert connection.get_api_key() == "project-secret"
    assert TopvisorProjectMapping.objects.filter(pk=selected.pk).exists()


def test_topvisor_credential_admin_is_read_only_and_has_no_add_button(
    client, django_user_model, settings
):
    project = Project.objects.create(name="Admin", domain="admin.example")
    connection = _connection(project, settings)
    admin_user = django_user_model.objects.create_superuser("admin", "admin@example.com", "secret")
    client.force_login(admin_user)

    changelist = client.get(reverse("admin:topvisor_topvisorcredential_changelist"))
    add_response = client.get(reverse("admin:topvisor_topvisorcredential_add"))
    detail = client.get(reverse("admin:topvisor_topvisorcredential_change", args=[connection.pk]))

    assert changelist.status_code == 200
    assert "Добавить реквизиты Topvisor" not in changelist.content.decode()
    assert add_response.status_code == 403
    assert detail.status_code == 200
    assert 'name="_save"' not in detail.content.decode()


def test_topvisor_mutations_require_login_post_and_csrf(client, django_user_model):
    project = Project.objects.create(name="Protected", domain="protected.example")
    connection_url = reverse("topvisor:connection", args=[project.id])
    settings_url = reverse("topvisor:credentials")
    assert client.post(connection_url, {}).status_code == 302
    assert client.post(settings_url, {}).status_code == 302
    user = django_user_model.objects.create_user("csrf", password="secret", is_staff=True)
    from django.test import Client

    csrf_client = Client(enforce_csrf_checks=True)
    csrf_client.force_login(user)
    assert csrf_client.post(settings_url, {"action": "credentials"}).status_code == 403
    assert csrf_client.post(settings_url, {"action": "delete"}).status_code == 403


def _connection(project, settings, user_id="uid", api_key="project-secret"):
    from apps.topvisor.models import TopvisorCredential

    settings.CREDENTIAL_ENCRYPTION_KEY = "stable-test-encryption-key"
    value = TopvisorCredential(user_id=user_id)
    value.set_api_key(api_key)
    value.last_verified_at = timezone.now()
    value.save()
    return value


@pytest.mark.parametrize(
    ("posted_user_id", "posted_api_key"),
    [("changed-user", ""), ("uid", "changed-api-key")],
)
def test_changed_credentials_remove_all_existing_mappings(
    client, settings, monkeypatch, posted_user_id, posted_api_key
):
    project = Project.objects.create(name="Changed", domain=f"{posted_user_id}.example")
    _connection(project, settings)
    selected = mapping(project)
    other_project = Project.objects.create(name="Other", domain=f"other-{posted_user_id}.example")
    other_selected = mapping(other_project)
    user = get_user_model().objects.create_user(posted_user_id, password="secret", is_staff=True)
    client.force_login(user)
    monkeypatch.setattr(TopvisorClient, "check_access", lambda self: ())

    response = client.post(
        reverse("topvisor:credentials"),
        {"action": "credentials", "user_id": posted_user_id, "api_key": posted_api_key},
    )
    assert response.status_code == 302
    assert not TopvisorProjectMapping.objects.filter(pk=selected.pk).exists()
    assert not TopvisorProjectMapping.objects.filter(pk=other_selected.pk).exists()


def test_unchanged_credentials_keep_existing_mapping(client, settings, monkeypatch):
    project = Project.objects.create(name="Unchanged", domain="unchanged.example")
    _connection(project, settings)
    selected = mapping(project)
    user = get_user_model().objects.create_user("unchanged", password="secret", is_staff=True)
    client.force_login(user)
    monkeypatch.setattr(TopvisorClient, "check_access", lambda self: ())
    response = client.post(
        reverse("topvisor:credentials"),
        {"action": "credentials", "user_id": "uid", "api_key": ""},
    )
    assert response.status_code == 302
    assert TopvisorProjectMapping.objects.filter(pk=selected.pk).exists()


def test_delete_credentials_removes_mappings_but_keeps_report_snapshot(
    client, settings, monkeypatch
):
    project = Project.objects.create(name="Delete", domain="delete-safe.example")
    _connection(project, settings)
    selected = mapping(project)
    report = Report.objects.create(project=project, report_month=date(2026, 7, 1))
    version = create_report_version(report=report)
    snapshot_id = version.snapshot.pk
    user = get_user_model().objects.create_user("delete-safe", password="secret", is_staff=True)
    client.force_login(user)
    settings.TOPVISOR_USER_ID, settings.TOPVISOR_API_KEY = "legacy", "legacy-secret"
    monkeypatch.setattr(TopvisorClient, "iter_projects", lambda self: iter([]))

    client.post(reverse("topvisor:credentials"), {"action": "delete"})
    assert Project.objects.filter(pk=project.pk).exists()
    assert Report.objects.filter(pk=report.pk).exists()
    assert type(version).objects.filter(pk=version.pk).exists()
    assert type(version.snapshot).objects.filter(pk=snapshot_id).exists()
    assert not TopvisorProjectMapping.objects.filter(pk=selected.pk).exists()
    html = client.get(reverse("topvisor:connection", args=[project.id])).content.decode()
    assert "Временный fallback" in html
    assert "Safe project" not in html


def test_encryption_failure_never_creates_pending_connection(client, settings, monkeypatch):
    from apps.topvisor.models import TopvisorCredential
    from apps.yandex.crypto import CredentialConfigurationError

    settings.CREDENTIAL_ENCRYPTION_KEY = "stable-test-encryption-key"
    user = get_user_model().objects.create_user("encrypt-fail", password="secret", is_staff=True)
    client.force_login(user)
    monkeypatch.setattr(TopvisorClient, "check_access", lambda self: ())
    monkeypatch.setattr(
        TopvisorCredential,
        "set_api_key",
        lambda self, key: (_ for _ in ()).throw(CredentialConfigurationError("api-secret")),
    )
    response = client.post(
        reverse("topvisor:credentials"),
        {"action": "credentials", "user_id": "uid", "api_key": "api-secret"},
    )
    assert response.status_code == 200
    assert "api-secret" not in response.content.decode()
    assert not TopvisorCredential.objects.exists()
    assert not TopvisorCredential.objects.filter(api_key_encrypted=b"pending").exists()


def test_unreadable_project_credentials_are_safe_on_page_and_sync(client, settings):
    from apps.topvisor.models import TopvisorCredential, TopvisorSyncRun

    project = Project.objects.create(name="Corrupt", domain="corrupt.example")
    connection = _connection(project, settings, api_key="never-display-this-secret")
    selected = mapping(project)
    TopvisorCredential.objects.filter(pk=connection.pk).update(api_key_encrypted=b"corrupt")
    settings.CREDENTIAL_ENCRYPTION_KEY = "different-encryption-key"
    user = get_user_model().objects.create_user("corrupt", password="secret")
    client.force_login(user)
    url = reverse("topvisor:connection", args=[project.id])

    response = client.get(url)
    html = response.content.decode()
    assert response.status_code == 200
    assert "Не удалось прочитать сохранённые реквизиты" in html
    assert "never-display-this-secret" not in html
    response = client.post(reverse("topvisor:sync", args=[project.id]), {"month": "2026-07"})
    assert response.status_code == 302
    run = TopvisorSyncRun.objects.get(mapping=selected)
    assert run.status == run.Status.FAILED
    assert "Не удалось прочитать сохранённые реквизиты" in run.error_message
    assert "corrupt" not in run.error_message


def test_global_credentials_are_shared_and_override_legacy(settings):
    from apps.topvisor.client import credentials_for_project

    settings.TOPVISOR_USER_ID, settings.TOPVISOR_API_KEY = "legacy-user", "legacy-key"
    first = Project.objects.create(name="First credentials", domain="first-credentials.example")
    second = Project.objects.create(name="Second credentials", domain="second-credentials.example")
    _connection(first, settings, "shared-user", "shared-key")

    first_credentials, first_legacy = credentials_for_project(first)
    second_credentials, second_legacy = credentials_for_project(second)
    assert (first_credentials.user_id, first_credentials.api_key, first_legacy) == (
        "shared-user",
        "shared-key",
        False,
    )
    assert (second_credentials.user_id, second_credentials.api_key, second_legacy) == (
        "shared-user",
        "shared-key",
        False,
    )


def test_history_rejects_request_without_dates_or_range():
    api = TopvisorClient(credentials=TopvisorCredentials("user", "secret"))
    with pytest.raises(ValueError, match="requires dates"):
        next(api.get_position_history("22653133", fields=["name"], positions_fields=["position"]))


class RealHistoryClient:
    def __init__(self, *, frequency=0, omit_frequency=False, fail_last=False):
        self.calls = []
        self.frequency = frequency
        self.omit_frequency = omit_frequency
        self.fail_last = fail_last

    def get_existing_position_dates(self, project_id, **params):
        self.calls.append(("dates", project_id, params))
        return ("2026-06-30", "2026-07-15", "2026-07-31")

    def get_position_history(self, project_id, **params):
        self.calls.append(("history", project_id, params))
        dates = params["dates"]
        volume = "volume:213:1:1"
        keyword = {
            "name": "seo",
            "positionsData": {
                f"{value}:{project_id}:2": {"position": "--" if value == dates[0] else 7}
                for value in dates
            },
        }
        if not self.omit_frequency:
            keyword[volume] = self.frequency
        yield {
            "headers": {"dates": dates, "projects": [{"id": project_id}]},
            "existsDates": dates,
            "keywords": [keyword],
        }
        if self.fail_last:
            raise TopvisorTemporaryError("safe")


def history_mapping(project):
    return TopvisorProjectMapping.objects.create(
        project=project,
        topvisor_project_id="22653133",
        selected_configurations=[
            {
                "id": "google:moscow",
                "searcher_name": "Google",
                "searcher_key": 1,
                "region_name": "Москва",
                "region_key": 213,
                "region_index": 2,
                "raw_depth": 2,
                "normalized_depth": 20,
            }
        ],
    )


def test_history_sync_uses_dates_and_normalizes_zero_and_missing_position():
    project = Project.objects.create(name="History", domain="history.example")
    api = RealHistoryClient(frequency=0)
    run = sync_positions(mapping=history_mapping(project), client=api)
    assert run.status == run.Status.SUCCESS
    calls = [call for call in api.calls if call[0] == "history"]
    assert calls and all(call[2].get("dates") for call in calls)
    assert all(call[1] == "22653133" for call in calls)
    assert all(call[2]["regions_indexes"] == ["2"] for call in calls)
    assert RankingSnapshot.objects.count() == 3
    assert set(KeywordPosition.objects.values_list("frequency", flat=True)) == {1}
    missing = KeywordPosition.objects.get(ranking_snapshot__snapshot_date=date(2026, 6, 30))
    assert missing.position_value is None
    assert missing.position_raw == "--"


def test_missing_frequency_and_last_page_failure_are_atomic():
    project = Project.objects.create(name="Atomic history", domain="atomic-history.example")
    selected = history_mapping(project)
    missing = sync_positions(mapping=selected, client=RealHistoryClient(omit_frequency=True))
    assert missing.status == missing.Status.FAILED
    assert "частотности" in missing.error_message
    assert not RankingSnapshot.objects.exists()
    failed = sync_positions(mapping=selected, client=RealHistoryClient(fail_last=True))
    assert failed.status == failed.Status.FAILED
    assert not RankingSnapshot.objects.exists()


@pytest.mark.parametrize("frequency", ["", -1, "invalid"])
def test_invalid_topvisor_frequency_is_atomic(frequency):
    project = Project.objects.create(
        name="Invalid frequency", domain=f"invalid-{str(frequency)}.example"
    )
    run = sync_positions(
        mapping=history_mapping(project), client=RealHistoryClient(frequency=frequency)
    )
    assert run.status == run.Status.FAILED
    assert "частотн" in run.error_message
    assert not RankingSnapshot.objects.exists()


def test_history_paginates_with_offset_and_keeps_required_parameters(monkeypatch):
    api = TopvisorClient(credentials=TopvisorCredentials("user", "secret"))
    calls = []

    def request(method, params):
        calls.append((method, params))
        count = 2 if params["offset"] == 0 else 1
        return {"keywords": [{"name": str(i)} for i in range(count)]}

    monkeypatch.setattr(api, "_request", request)
    pages = list(
        api.get_position_history(
            22653133,
            page_size=2,
            dates=["2026-07-01", "2026-07-02"],
            regions_indexes=["2"],
            fields=["name"],
            positions_fields=["position"],
        )
    )
    assert len(pages) == 2
    assert [params["offset"] for _method, params in calls] == [0, 2]
    for method, params in calls:
        assert method == "get/positions_2/history"
        assert params["project_id"] == "22653133"
        assert params["dates"]
        assert params["regions_indexes"] == ["2"]
        assert params["fields"] and params["positions_fields"]
        assert params["show_headers"] == params["show_exists_dates"] == 1
        assert params["limit"] == 2


def test_sync_batches_existing_dates():
    project = Project.objects.create(name="Batches", domain="batches.example")
    selected = history_mapping(project)

    class BatchClient(RealHistoryClient):
        def get_existing_position_dates(self, project_id, **params):
            return tuple(f"2026-07-{day:02d}" for day in range(1, 22))

    api = BatchClient(frequency=1)
    run = sync_positions(mapping=selected, client=api)
    assert run.status == run.Status.SUCCESS
    batches = [params["dates"] for kind, _project, params in api.calls if kind == "history"]
    assert [len(batch) for batch in batches] == [20, 1]
    assert RankingSnapshot.objects.count() == 21


def test_yandex_frequency_alias_is_shared_with_google_and_missing_is_atomic():
    project = Project.objects.create(name="Shared volume", domain="shared-volume.example")
    selected = TopvisorProjectMapping.objects.create(
        project=project,
        topvisor_project_id="22653133",
        selected_configurations=[
            {
                "id": "yandex",
                "searcher_name": "Yandex",
                "searcher_key": 0,
                "region_name": "Москва",
                "region_key": 213,
                "region_index": 1,
                "raw_depth": 1,
                "normalized_depth": 100,
            },
            {
                "id": "google",
                "searcher_name": "Google",
                "searcher_key": 1,
                "region_name": "Москва",
                "region_key": 213,
                "region_index": 2,
                "raw_depth": 2,
                "normalized_depth": 20,
            },
        ],
    )

    class SharedVolumeClient:
        missing = False

        def get_existing_position_dates(self, project_id, **params):
            return ("2026-07-31",)

        def get_position_history(self, project_id, **params):
            region = params["regions_indexes"][0]
            keywords = [
                {
                    "name": "zero query",
                    "frequency_alias": 0,
                    "positionsData": {f"2026-07-31:{project_id}:{region}": {"position": 3}},
                },
                {
                    "name": "positive query",
                    "frequency_alias": 42,
                    "positionsData": {f"2026-07-31:{project_id}:{region}": {"position": 7}},
                },
            ]
            if self.missing:
                keywords[1].pop("frequency_alias")
            yield {
                "headers": {
                    "dates": ["2026-07-31"],
                    "fields": [{"name": "volume:213:0:1", "alias": "frequency_alias"}],
                    "fieldsLabels": {"frequency_alias": "volume:213:0:1"},
                },
                "existsDates": ["2026-07-31"],
                "keywords": keywords,
            }

    api = SharedVolumeClient()
    run = sync_positions(mapping=selected, client=api)
    assert run.status == run.Status.SUCCESS
    assert RankingSnapshot.objects.count() == 2
    for snapshot in RankingSnapshot.objects.all():
        assert dict(snapshot.positions.values_list("normalized_query", "frequency")) == {
            "zero query": 1,
            "positive query": 42,
        }

    RankingSnapshot.objects.all().delete()
    api.missing = True
    failed = sync_positions(mapping=selected, client=api)
    assert failed.status == failed.Status.FAILED
    assert "1" in failed.error_message
    assert "positive query" not in failed.error_message
    assert not RankingSnapshot.objects.exists()
