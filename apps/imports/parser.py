import csv
import math
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from itertools import chain
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook

from apps.metrics.models import KeywordPosition
from apps.metrics.normalization import normalize_frequency

MAX_HEADER_SCAN_ROWS = 20
MAX_XLSX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024

HEADER_ALIASES = {
    "query": {
        "query",
        "keyword",
        "phrase",
        "search query",
        "запрос",
        "ключевая фраза",
        "ключевой запрос",
        "фраза",
    },
    "position": {"position", "rank", "позиция", "место"},
    "frequency": {
        "frequency",
        "search volume",
        "volume",
        "частотность",
        "частота",
    },
    "group_name": {"group", "group name", "группа", "группа запросов"},
    "target_url": {
        "target url",
        "landing page",
        "url",
        "релевантная страница",
        "целевая страница",
    },
}


class ImportFileError(Exception):
    pass


@dataclass(frozen=True)
class ParsedError:
    row_number: int
    code: str
    message: str
    raw_values: dict[str, str]


@dataclass(frozen=True)
class ImportPreview:
    total_rows: int
    valid_rows: list[dict]
    errors: list[ParsedError]

    @property
    def error_row_count(self):
        return len({error.row_number for error in self.errors})


def parse_position_file(
    filename: str, data: bytes, snapshot_date: date | None = None
) -> ImportPreview:
    extension = Path(filename).suffix.casefold()
    if extension == ".csv":
        rows = _read_csv(data)
    elif extension == ".xlsx":
        rows = _read_xlsx(data)
    else:
        raise ImportFileError("Поддерживаются только файлы CSV и XLSX.")
    return _parse_rows(rows, snapshot_date=snapshot_date)


def _read_csv(data: bytes):
    if b"\x00" in data:
        raise ImportFileError("CSV содержит недопустимые нулевые байты.")

    text = None
    for encoding in ("utf-8-sig", "cp1251"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ImportFileError("Не удалось определить кодировку CSV. Используйте UTF-8 или CP1251.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        return csv.reader(StringIO(text), delimiter=";")
    return csv.reader(StringIO(text), dialect)


def _read_xlsx(data: bytes):
    _validate_xlsx_archive(data)
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ImportFileError("Не удалось прочитать XLSX. Проверьте структуру файла.") from exc

    worksheet = workbook.active
    return worksheet.iter_rows(values_only=True)


def _validate_xlsx_archive(data: bytes):
    if not data.startswith(b"PK"):
        raise ImportFileError("Файл с расширением XLSX не является Excel-документом.")
    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            names = {item.filename for item in archive.infolist()}
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise ImportFileError("XLSX не содержит обязательных частей книги Excel.")

            total_size = 0
            for item in archive.infolist():
                path_parts = Path(item.filename).parts
                if item.filename.startswith("/") or ".." in path_parts:
                    raise ImportFileError("XLSX содержит небезопасные пути внутри архива.")
                if item.flag_bits & 0x1:
                    raise ImportFileError("Зашифрованные XLSX не поддерживаются.")
                total_size += item.file_size
                if total_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ImportFileError("Распакованный XLSX слишком большой.")
                if item.file_size > max(item.compress_size, 1) * 1000:
                    raise ImportFileError("XLSX имеет подозрительно высокий коэффициент сжатия.")
    except zipfile.BadZipFile as exc:
        raise ImportFileError("Поврежденный XLSX-файл.") from exc


def _parse_rows(rows, snapshot_date=None) -> ImportPreview:
    row_iterator = enumerate(rows, start=1)
    buffered_rows = []
    header_row_number = None
    headers = None
    column_map = None

    for row_number, row in row_iterator:
        row = list(row)
        buffered_rows.append((row_number, row))
        candidate_map = _build_column_map(row, snapshot_date=snapshot_date)
        if {"query", "position", "frequency"}.issubset(candidate_map):
            header_row_number = row_number
            headers = [_string_value(value) for value in row]
            column_map = candidate_map
            break
        if len(buffered_rows) >= MAX_HEADER_SCAN_ROWS:
            break

    if header_row_number is None or headers is None or column_map is None:
        raise ImportFileError(
            "Не найдены обязательные столбцы «Запрос», «Позиция» и «Частотность» "
            "в первых 20 строках."
        )

    remaining_rows = chain(
        ((number, row) for number, row in buffered_rows if number > header_row_number),
        row_iterator,
    )
    valid_rows = []
    errors = []
    seen_keys = set()
    total_rows = 0

    for row_number, raw_row in remaining_rows:
        row = list(raw_row)
        if not any(_string_value(value) for value in row):
            continue
        total_rows += 1
        raw_values = _raw_values(headers, row)
        parsed, row_errors = _parse_data_row(row_number, row, column_map, raw_values)
        if row_errors:
            errors.extend(row_errors)
            continue

        duplicate_key = (parsed["normalized_query"], parsed["group_name"].casefold())
        if duplicate_key in seen_keys:
            errors.append(
                ParsedError(
                    row_number,
                    "duplicate_query",
                    "Запрос с такой группой уже встречался в файле.",
                    raw_values,
                )
            )
            continue
        seen_keys.add(duplicate_key)
        valid_rows.append(parsed)

    if total_rows == 0:
        raise ImportFileError("В файле нет строк с позициями.")
    return ImportPreview(total_rows, valid_rows, errors)


def _build_column_map(row, snapshot_date=None):
    result = {}
    for index, value in enumerate(row):
        normalized = _normalize_header(value)
        if not normalized:
            continue
        for field_name, aliases in HEADER_ALIASES.items():
            if normalized in aliases:
                result.setdefault(field_name, index)
                break
        else:
            if normalized.startswith(("позиция ", "position ")):
                result.setdefault("position", index)
            elif normalized.startswith(("частотность ", "частота ", "frequency ")):
                result.setdefault("frequency", index)
            elif normalized.startswith(("запрос ", "ключевая фраза ", "keyword ")):
                result.setdefault("query", index)
            elif normalized.startswith(("группа ", "group ")):
                result.setdefault("group_name", index)
            elif normalized.startswith(("релевантная страница ", "target url ")):
                result.setdefault("target_url", index)
        if snapshot_date and _date_from_header(value) == snapshot_date:
            result.setdefault("position", index)
    return result


def _date_from_header(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _string_value(value)
    for pattern in (r"\b(\d{4})-(\d{2})-(\d{2})\b", r"\b(\d{2})[./](\d{2})[./](\d{4})\b"):
        match = re.search(pattern, text)
        if not match:
            continue
        try:
            if pattern.startswith(r"\b(\d{4})"):
                return date(int(match[1]), int(match[2]), int(match[3]))
            return date(int(match[3]), int(match[2]), int(match[1]))
        except ValueError:
            return None
    return None


def _parse_data_row(row_number, row, column_map, raw_values):
    errors = []
    query = _cell(row, column_map["query"])
    query = " ".join(_string_value(query).split())
    if not query:
        errors.append(ParsedError(row_number, "missing_query", "Не указан запрос.", raw_values))
    elif len(query) > 500:
        errors.append(
            ParsedError(row_number, "query_too_long", "Запрос длиннее 500 символов.", raw_values)
        )

    position_raw = _string_value(_cell(row, column_map["position"]))
    if len(position_raw) > 50:
        errors.append(
            ParsedError(row_number, "position_too_long", "Позиция длиннее 50 символов.", raw_values)
        )
    try:
        position_value, position_status = _parse_position(position_raw)
    except ValueError as exc:
        errors.append(ParsedError(row_number, "invalid_position", str(exc), raw_values))
        position_value = None
        position_status = KeywordPosition.Status.NOT_FOUND

    frequency_raw = _cell(row, column_map["frequency"])
    try:
        frequency = _parse_frequency(frequency_raw)
    except ValueError as exc:
        errors.append(ParsedError(row_number, "invalid_frequency", str(exc), raw_values))
        frequency = 0

    group_name = ""
    if "group_name" in column_map:
        group_name = " ".join(_string_value(_cell(row, column_map["group_name"])).split())
        if len(group_name) > 255:
            errors.append(
                ParsedError(
                    row_number,
                    "group_too_long",
                    "Название группы длиннее 255 символов.",
                    raw_values,
                )
            )

    target_url = ""
    normalized_target_url = ""
    if "target_url" in column_map:
        target_url = _string_value(_cell(row, column_map["target_url"])).strip()
        if len(target_url) > 2000:
            errors.append(
                ParsedError(row_number, "url_too_long", "URL длиннее 2000 символов.", raw_values)
            )
        elif target_url:
            try:
                normalized_target_url = normalize_url(target_url)
            except ValueError as exc:
                errors.append(ParsedError(row_number, "invalid_url", str(exc), raw_values))

    parsed = {
        "query": query,
        "normalized_query": normalize_query(query),
        "frequency": frequency,
        "position_raw": position_raw,
        "position_value": position_value,
        "position_status": position_status,
        "group_name": group_name,
        "target_url": target_url,
        "normalized_target_url": normalized_target_url,
    }
    return parsed, errors


def normalize_query(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value)
    return " ".join(normalized.casefold().split())


def normalize_url(value: str) -> str:
    value = unicodedata.normalize("NFKC", value.strip())
    if value.startswith("/") and not value.startswith("//"):
        parsed = urlsplit(value)
        return urlunsplit(("", "", parsed.path or "/", parsed.query, ""))
    if value.startswith("//"):
        value = f"https:{value}"
    elif "://" not in value:
        value = f"https://{value}"

    try:
        parsed = urlsplit(value)
        if parsed.scheme.casefold() not in {"http", "https"} or not parsed.hostname:
            raise ValueError
        hostname = parsed.hostname.rstrip(".").casefold()
        if hostname.startswith("www."):
            hostname = hostname[4:]
        hostname = hostname.encode("idna").decode("ascii")
        port = parsed.port
    except (UnicodeError, ValueError) as exc:
        raise ValueError("Некорректный URL.") from exc

    default_port = (parsed.scheme.casefold() == "http" and port == 80) or (
        parsed.scheme.casefold() == "https" and port == 443
    )
    netloc = hostname if port is None or default_port else f"{hostname}:{port}"
    return urlunsplit((parsed.scheme.casefold(), netloc, parsed.path or "/", parsed.query, ""))


def _parse_position(value: str):
    normalized = value.strip().casefold().replace(" ", "")
    if normalized in {"", "-", "—", "нет", "не найден", "notfound", "n/a", "na"}:
        return None, KeywordPosition.Status.NOT_FOUND
    if normalized in {">100", "100+", "101+", "за100"}:
        return None, KeywordPosition.Status.BEYOND_100

    normalized = normalized.replace(",", ".")
    try:
        numeric = float(normalized)
    except ValueError as exc:
        raise ValueError("Позиция должна быть числом, «>100» или «не найден».") from exc
    if not math.isfinite(numeric) or not numeric.is_integer() or numeric < 1:
        raise ValueError("Позиция должна быть целым положительным числом.")
    position = int(numeric)
    if position > 100:
        return None, KeywordPosition.Status.BEYOND_100
    return position, KeywordPosition.Status.RANKED


def _parse_frequency(value):
    try:
        return normalize_frequency(value)
    except ValueError as exc:
        if not _string_value(value).strip():
            raise ValueError("Частотность обязательна для каждой строки.") from exc
        raise ValueError("Частотность должна быть целым неотрицательным числом.") from exc


def _normalize_header(value):
    value = unicodedata.normalize("NFKC", _string_value(value)).casefold().replace("ё", "е")
    value = re.sub(r"[^a-zа-я0-9]+", " ", value)
    return " ".join(value.split())


def _raw_values(headers, row):
    result = {}
    for index, header in enumerate(headers):
        key = header or f"column_{index + 1}"
        result[key] = _string_value(_cell(row, index))
    return result


def _cell(row, index):
    return row[index] if index < len(row) else None


def _string_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
